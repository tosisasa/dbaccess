#Prassデータベース照会のサンプルプログラム

import tkinter
from tkinter import messagebox
import openpyxl as excel
import DBAccess

# OKボタンクリック
def ok_click():
    # テキストボックスの内容を得る
    sColumn3 = text1.get()

    if sColumn3 == "":
        messagebox.showinfo("ERROR", "必須入力です。")
        return


    sql = "SELECT 'COLUMN1', 'COLUMN2' FROM DUAL WHERE 'xxx' = '" + sColumn3 + "'"
    rows = DBAccess.selectDB(sql, "ユーザ名@ホスト名")

    out.delete(0, tkinter.END) #SQLログ出力
    out.insert(tkinter.END, sql)

    if len(rows) > 0:
        # 新規ワークブックを作ってシートを得る --- (*2)
        wb = excel.Workbook()
        ws = wb.active

        ws["A1"] = "カラム1"
        ws["B1"] = "カラム2"

        iRow=2
        for r in rows:

            ws.cell(column=1, row=iRow, value=r[0])
            ws.cell(column=2, row=iRow, value=r[1])
            iRow = iRow + 1

        wb.save("result.xlsx")

        messagebox.showinfo("", "Excelに出力しました")

    
    else:
        messagebox.showinfo("ERROR", "データがありません。")


# ウィンドウを作成
root = tkinter.Tk()
root.title("＊＊＊＊照会")
root.geometry("500x250") # サイズを指定

# 入力１
label1 = tkinter.Label(text="カラム1")
label1.grid(row=1, column=1, padx=10,)

text1 = tkinter.Entry(width=40)
text1.grid(row=1, column=2)

# 入力２
label2 = tkinter.Label(text="カラム2")
label2.grid(row=2, column=1, padx=10,)

text2 = tkinter.Entry(width=40)
text2.grid(row=2, column=2)

# ボタン
okButton = tkinter.Button(text='OK', command=ok_click)
#okButton.place(x=10, y=80)
okButton.grid(row=4, column=3)

#出力
out = tkinter.Entry(width=40)
out.grid(row=6, column=2)


# ウィンドウを動かす
root.mainloop()
